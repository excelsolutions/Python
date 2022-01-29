# 2021-01-16 moving all widgets to frame
# 2021-01-08 V1.4
from tkinter.messagebox import showinfo

from PIL import Image, ImageTk
import PIL
import os
import glob
import math
import \
    tkinter as tk
# link: https://stackoverflow.com/questions/17466561/best-way-to-strucd yhn6ture-a-tkinter-application
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import tkinter.ttk as ttk
from tkinter import filedialog
import shutil  # to files copy


class MainApplication(tk.Frame):
    def __init__(self, parent, *args, **kwargs):
        tk.Frame.__init__(self, parent, *args, **kwargs)
        self.parent = parent
        # Style
        font_main = ("Courier", 14)
        self.style_Table = ttk.Style()
        self.style_Table.configure("Treeview",
                                   font=(None, 10),
                                   background="gray99",
                                   foreground="black",
                                   rowheight=20,
                                   fieldbackground="silver")
        self.style_Table.map('Treeview', background=[('selected', 'green')])
        # Read settings from xlsx
        self.wb = load_workbook("Photo_To_Onedrive.xlsx")
        self.sh_settings = self.wb['settings']
        path_folder = self.sh_settings['B1'].value
        path_onedrive_folder = self.sh_settings['B2'].value

        # MAIN FRAMES
        self.frame_Main_Top = tk.LabelFrame(root, text='Settings', width=root.winfo_width())
        self.frame_Main_Top.pack(side='top', anchor='nw', expand=False)
        self.frame_Main_Bottom = tk.LabelFrame(root, text="Files")
        self.frame_Main_Bottom.pack(side='top', anchor='nw', expand=False)
        self.frame_Main_Right = tk.Frame(self.frame_Main_Bottom)
        self.frame_Main_Right.pack(side='right', fill='both', anchor='ne')
        self.frame_Folders = tk.Frame(self.frame_Main_Top)
        self.frame_Folders.grid(row=0, column=0)

        self.lbl_Title = tk.Label(self.frame_Main_Top, text="Pictures to Onedrive", bg='yellow', font=font_main)
        self.lbl_Title.grid(row=0, column=0, sticky='w')
        # FRAME FOLDER WITH IMAGES
        self.frame_Folder = ttk.LabelFrame(self.frame_Main_Top, text="Folder with images")
        self.frame_Folder.grid(row=1, column=0, sticky='w')
        self.txt_Folder = tk.Entry(self.frame_Folder, font=font_main, width=60)
        self.txt_Folder.grid(row=0, column=0, sticky='w')
        try:
            if path_folder:
                self.txt_Folder.delete(0, tk.END)
                self.txt_Folder.insert(0, path_folder)
        except:
            showinfo(title='Information', message="Problem!")

        self.btn_Folder = tk.Button(self.frame_Folder, text='Pick folder', command=self.pick_folder, font=font_main)
        self.btn_Folder.grid(row=0, column=1)
        self.btn_Folder = tk.Button(self.frame_Folder, text='Run', command=self.proceed_Files, font=font_main)
        self.btn_Folder.grid(row=0, column=2)
        # FRAME ONEDRIVE FOLDER
        self.frame_Onedrive_Folder = tk.LabelFrame(self.frame_Main_Top, text="Path to Onedrive folder")
        self.frame_Onedrive_Folder.grid(row=2, column=0)
        self.txt_Onedrive_Folder = tk.Entry(self.frame_Onedrive_Folder, font=font_main, width=60)
        self.txt_Onedrive_Folder.grid(row=0, column=0)
        self.txt_Onedrive_Folder.delete(0, tk.END)
        self.txt_Onedrive_Folder.insert(0, path_onedrive_folder)
        self.btn_Onedrive_Folder = tk.Button(self.frame_Onedrive_Folder,
                                             text='Pick Onedrive folder', command=self.pick_folder_onedrive,
                                             font=font_main)
        self.btn_Onedrive_Folder.grid(row=0, column=1)

        # STATISTIC
        self.frame_Top_Count = ttk.LabelFrame(self.frame_Main_Top, text='Statistic data')
        self.frame_Top_Count.grid(row=3, column=0, sticky='w')

        self.lbl_jpg = ttk.Label(self.frame_Top_Count, text='JPG')
        self.lbl_jpg.grid(row=0, column=0)
        self.lbl_jpg_e = ttk.Label(self.frame_Top_Count,
                                   text=self.files_counter(self.txt_Folder.get(), ('jpg', 'JPG', 'jpeg', 'JPEG')))
        self.lbl_jpg_e.grid(row=1, column=0)

        self.lbl_png = ttk.Label(self.frame_Top_Count, text='PNG')
        self.lbl_png.grid(row=0, column=1)
        self.lbl_png_e = ttk.Label(self.frame_Top_Count,
                                   text=self.files_counter(self.txt_Folder.get(), ('png', 'PNG')))
        self.lbl_png_e.grid(row=1, column=1)

        self.lbl_raw = ttk.Label(self.frame_Top_Count, text='RAW')
        self.lbl_raw.grid(row=0, column=2)
        self.lbl_raw_e = ttk.Label(self.frame_Top_Count,
                                   text=self.files_counter(self.txt_Folder.get(), ('RW2', 'NEF')))
        self.lbl_raw_e.grid(row=1, column=2)

        # TABLE Treeview
        # scrollbar
        self.frame_Treeview = tk.LabelFrame(self.frame_Main_Bottom)
        self.frame_Treeview.pack(side='left', anchor='ne')
        self.scroll_y = ttk.Scrollbar(self.frame_Treeview)
        self.scroll_y.pack(side='right', fill='y')
        self.table_Files = ttk.Treeview(self.frame_Treeview, yscrollcommand=self.scroll_y.set, height=200)

        self.scroll_y.config(command=self.table_Files.yview)
        self.table_Files['columns'] = ('No', 'Name of file', 'Width x Height', 'Compression level', 'File size')
        self.table_Files.column("#0", width=0, stretch=False)
        self.table_Files.column("No", width=10)
        self.table_Files.column("Name of file", width=150)
        self.table_Files.column("Width x Height", width=150)
        self.table_Files.column("Compression level", width=150)
        self.table_Files.column("File size", width=200)

        self.table_Files.heading("#0", text="")
        self.table_Files.heading("No", text="No")
        self.table_Files.heading("Name of file", text="Name of file")
        self.table_Files.heading("Width x Height", text="Width x Height")
        self.table_Files.heading("Compression level", text="Compression level")
        self.table_Files.heading("File size", text="File size")

        self.table_Files.bind('<<TreeviewSelect>>', self.item_selected)
        self.table_Files.pack(side='left', fill='both', anchor='nw')

        try:
            if self.txt_Folder.get():
                self.load_images(self.txt_Folder.get())
        except:
            showinfo(title='Information', message='Problem')

        # FRAME ACTIONS
        '''
        |----------|
        |       OOO|
        |          |
        |----------|
        '''
        self.action_Frame = tk.LabelFrame(self.frame_Main_Top, text="Prefix added to filename")
        self.action_Frame.grid(row=1, column=3)
        self.txt_Prefix = tk.Entry(self.action_Frame, font=font_main, width=20)
        self.txt_Prefix.grid(row=0, column=0)

        # Picture preview
        '''
        |----------|
        |          |
        |       OOO|
        |----------|
        '''
        self.preview_Frame = tk.LabelFrame(self.frame_Main_Right)
        self.preview_Frame.pack(fill='both', side='bottom', expand=True, anchor='sw')
        self.preview_Label = ttk.Label(self.preview_Frame, width=root.winfo_width())
        self.preview_Label.pack(fill='both', side='bottom')
        self.build_excel()

    def pick_folder_onedrive(self):
        file_path = os.getcwd()
        folder_path = filedialog.askdirectory()
        self.txt_Onedrive_Folder.delete(0, tk.END)
        self.txt_Onedrive_Folder.insert(0, folder_path)

        self.sh_settings['B2'].value = folder_path
        self.wb.save(filename=os.path.join(file_path, "Photo_To_Onedrive.xlsx"))

    def pick_folder(self):
        folder_path = filedialog.askdirectory()
        if os.path.isdir(folder_path):
            self.load_images(folder_path)

    def load_images(self, folder_path):
        file_path = os.getcwd()
        i = 0
        self.txt_Folder.delete(0, tk.END)
        self.txt_Folder.insert(0, folder_path)
        self.sh_settings['B1'].value = folder_path
        self.wb.save(filename=os.path.join(file_path, "Photo_To_Onedrive.xlsx"))
        for j in self.table_Files.get_children():
            self.table_Files.delete(j)
        files = os.listdir(folder_path)
        images = [file
                  for file in files
                  if file.endswith(('jpg', 'png', 'JPG', 'PNG'))]
        for image in images:
            i = i + 1
            img = PIL.Image.open(os.path.join(folder_path, image))
            width, height = img.size
            self.table_Files.insert(parent='',
                                    index='end',
                                    iid=i,
                                    text='',
                                    values=(i, image, '{} x {}'.format(width, height), 'NONE',
                                            '{:0.2f} MB'.format(
                                                os.path.getsize(os.path.join(folder_path, image)) / 1024 / 1024, 2)))

    def files_counter(self, path, extensions):
        i = 0
        files = os.listdir(path)
        s_files = [file
                   for file in files
                   if file.endswith(extensions)]
        for f in s_files:
            i = i + 1
        return i

    def copy_Files(self, source, destination):
        try:
            shutil.copyfile(source, destination)
            print("File copied successfully.")

        # If source and destination are same
        except shutil.SameFileError:
            print("Source and destination represents the same file.")

        # If destination is a directory.
        except IsADirectoryError:
            print("Destination is a directory.")

        # If there is any permission issue
        except PermissionError:
            print("Permission denied.")

        # For other errors
        except:
            print("Error occurred while copying file.")

    def proceed_Files(self):
        path = self.txt_Folder.get()
        without_extra_slash = os.path.normpath(path)
        folder = os.path.basename(without_extra_slash)

        wb_images = Workbook()
        ws = wb_images.active
        ws.title = "List of images"
        ws['A1'].value = 'LP.'
        ws['B1'].value = 'Filename'
        ws['C1'].value = 'Resolution'
        ws['D1'].value = 'Compression lvl'
        ws['E1'].value = 'Size'
        for item in self.table_Files.get_children():
            item = self.table_Files.item(item)
            record = item['values']
            filename = os.path.join(path, (record[1]))
            new_filename = os.path.join(path, ('zmn' + record[1]))
    def item_selected(self, event):
        for selected_item in self.table_Files.selection():
            item = self.table_Files.item(selected_item)
            record = item['values']
            # showinfo(title='Information', message=record[1])
        img_path = os.path.join(self.txt_Folder.get(), str(record[1]))

        image_resized = Image.open(img_path)
        width, height = image_resized.size
        new_width = root.winfo_width() - 600  # self.frame_Treeview.winfo_width()
        new_height = math.floor((new_width * height) / width)
        print(new_width, new_height)
        image_resized = image_resized.resize((new_width, new_height), Image.ANTIALIAS)
        img = ImageTk.PhotoImage(image_resized)

        self.preview_Label.config(image=img, text=img_path + chr(10) + str(record[1]))
        self.preview_Label.image = img

    def build_excel(self):
        wb_images = Workbook()
        ws = wb_images.active
        ws.title = "List of images"
        path = self.txt_Folder.get()
        without_extra_slash = os.path.normpath(path)
        folder = os.path.basename(without_extra_slash)
        i = 2
        ws['A1'].value = 'LP.'
        ws['B1'].value = 'Filename'
        ws['C1'].value = 'Resolution'
        ws['D1'].value = 'Compression lvl'
        ws['E1'].value = 'Size'
        for item in self.table_Files.get_children():
            item = self.table_Files.item(item)
            record = item['values']
            ws['A' + str(i)].value = str(record[0]) + "."
            ws['B' + str(i)].value = record[1]
            ws['C' + str(i)].value = record[2]
            ws['D' + str(i)].value = record[3]
            ws['E' + str(i)].value = record[4]

            i = i + 1


        wb_images.save(filename=os.path.join(path, (folder + ".xlsx")))



def compress_images(directory=False, quality=30):
    # 1. If there is a directory then change into it, else perform the next operations inside of the
    # current working directory:
    if directory:
        os.chdir(directory)

    # 2. Extract all of files
    files = os.listdir()

    # 3. Extract all of the images:
    images = [file for file in files if file.endswith(('jpg', 'png', 'JPG', 'PNG'))]

    # 4. Loop over every image:
    for image in images:
        print(image)

        # 5. Open every image:
        img = Image.open(image)

        # 5. Compress every image and save it with a new name:
        img.save("Kopia_" + image, optimize=True, quality=quality)


def compress_one_image(directory=False, quality=10):
    if directory:
        os.chdir(os.path.dirname(os.path.abspath(directory)))
    img = Image.open(directory)
    img.save('_' + os.path.basename(directory), optimize=True, quality=quality)


def images_change_resolution(directory=False, quality=50, resize=0.8, png_jpg_conv=False, file=None):
    if directory:
        os.chdir(directory)

    # 2. Extract all of files
    files = os.listdir()

    # 3. Extract all of the images:
    images = [file for file in files if file.endswith(('jpg', 'png', 'JPG', 'PNG'))]

    # 4. Loop over every image:
    for image in images:
        print(image)

        # 5. Open every image:
        img = Image.open(image)

        # 5. Compress every image and save it with a new name:
        width, height = img.size
        old_width, old_height = img.size
        width = math.floor(width * resize)
        height = math.floor(height * resize)
        img = img.resize((width, height))
        extension = os.path.splitext(file)[1]
        filename = os.path.basename(file)
        if (extension == '.png' and png_jpg_conv == True):
            img = img.convert('RGB')
            quality = 80
            filename = filename.replace('.png', '_zmn' + '.jpg')
        else:
            filename = filename.replace(extension, '_zmn' + extension)
        img.save(filename, optimize=True, quality=quality)


def image_change_resolution(file=False, quality=50, resize=0.5, png_jpg_conv=False):
    if file:
        os.chdir(os.path.dirname(os.path.abspath(file)))
    img = Image.open(file)
    width, height = img.size
    old_width, old_height = img.size
    width = math.floor(width * resize)
    height = math.floor(height * resize)
    img = img.resize((width, height))
    extension = os.path.splitext(file)[1]
    filename = os.path.basename(file)
    if (extension == '.png' and png_jpg_conv == True):
        img = img.convert('RGB')
        quality = 80
        filename = filename.replace('.png', '_zmn' + '.jpg')
    # else:
    #    filename = filename.replace(extension, '_zmn' + extension) # bez "zmn"
    img.save(filename, optimize=True, quality=quality)

    print('Resize img completed: [ratio = ' + str(resize) + ']')
    print('Width: ', old_width, ' -->', width)
    print('Height: ', old_height, '-->', height)


def image_info(directory=False):
    if directory:
        os.chdir(os.path.dirname(os.path.abspath(directory)))
    img = Image.open(directory)
    width, height = img.size
    print('width: ', width, 'height: ', height)


# subdirectory_path = r'G:\głoszenia'
# dir_path_one_image = r'G:\Dom_P1090417.JPG'
# compress_one_image(dir_path_one_image, 30)
# image_change_resolution(dir_path_one_image, 30, 0.6, True)
# image_info(dir_path_one_image)
# images_change_resolution(subdirectory_path, resize=0.5)
# compress_images(directory=subdirectory_path)

# for filename in os.listdir(subdirectory_path):
#      sciezka= os.path.join(subdirectory_path, filename)
#      img = Image.open(sciezka)
#      width, height = img.size
#      if width>2000 or height>2000:
#         print(filename, width, height)
#         image_change_resolution(sciezka,quality= 50, resize=0.5,png_jpg_conv= True)

if __name__ == '__main__':
    root = tk.Tk()
    root.title('Pictures to Onedrive')
    MainApplication(root).pack(side="top", fill="both", expand=True)
    root.geometry('1500x1000')
    root.mainloop()
